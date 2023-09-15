# -*- coding: utf-8 -*-
# Copyright 2019 Artem Shurshilov
# Odoo Proprietary License v1.0

# This software and associated files (the "Software") may only be used (executed,
# modified, executed after modifications) if you have purchased a valid license
# from the authors, typically via Odoo Apps, or if you have received a written
# agreement from the authors of the Software (see the COPYRIGHT file).

# You may develop Odoo modules that use the Software as a library (typically
# by depending on it, importing it and using its resources), but without copying
# any source code or material from the Software. You may distribute those
# modules under the license of your choice, provided that this license is
# compatible with the terms of the Odoo Proprietary License (For example:
# LGPL, MIT, or proprietary licenses similar to this one).

# It is forbidden to publish, distribute, sublicense, or sell copies of the Software
# or modified copies of the Software.

# The above copyright notice and this permission notice must be included in all
# copies or substantial portions of the Software.

# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT.
# IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM,
# DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE,
# ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER
# DEALINGS IN THE SOFTWARE.

import base64
import io
from odoo import models, fields, api, _
import openpyxl
from odoo.exceptions import UserError
import os
import sys
import subprocess
import datetime
import re
import logging
import base64
_logger = logging.getLogger(__name__)


class IrActionsReport(models.Model):
    _inherit = 'ir.actions.report'

    def _get_path_source_excel(self):
        if sys.platform == 'win32':
            return r'C:\temp\backups\out.xlsx'
        return '/tmp/backups/out.xlsx'

    def _get_path_convert_excel(self):
        if sys.platform == 'win32':
            return r'C:\temp\backups'
        return '/tmp/backups'

    def _get_path_libreoffice_excel(self):
        # TODO: Provide support for more platforms
        if sys.platform == 'darwin':
            return '/Applications/LibreOffice.app/Contents/MacOS/soffice'
        if sys.platform == 'win32':
            return "C:\Program Files\LibreOffice\program\soffice.exe"
        return 'libreoffice'

    report_type = fields.Selection(
        selection_add=[('excel', 'EXCEL')], ondelete={'excel': 'cascade'})
    template_excel = fields.Binary(string='Excel template', attachment=True)
    excel_out_report_type = fields.Selection(
        selection=[('excel', 'EXCEL'), ('pdf', 'PDF'),
                   ('ods', 'ODS'), ('odt', 'ODT'), ('html', 'HTML')],
        help="Extension will download users", default='excel')
    excel_path_source = fields.Char(
        string='OS path to source file temporary', default=_get_path_source_excel)
    excel_path_convert_folder = fields.Char(
        string='OS path to converted file temporary', default=_get_path_convert_excel)
    excel_path_libreoffice = fields.Char(
        string='OS path to libreoffice', default=_get_path_libreoffice_excel, help="For linux just libreoffice")

    def _render_excel(self, report_ref, res_ids, data=None):
        return self.render_excel(report_ref, res_ids, data)

    @api.model
    def render_excel(self, report_ref, docids, data=None):
        if not data:
            data = {}
        data.setdefault('report_type', 'excel')
        report = self._get_report(report_ref)
        data = self._get_rendering_context(report, docids, data)
        # READ DATA
        content = base64.b64decode(self.template_excel)

        # MERGE DATA
        # open xcel sheets
        wb1 = openpyxl.load_workbook(io.BytesIO(content))
        ws1 = wb1.active

        # compare each element
        for doc in data['docs']:
            for row in range(ws1.max_row):
                for column in range(ws1.max_column):
                    val = ws1.cell(row=row+1, column=column+1).value
                    if isinstance(val, str):
                        result = re.findall(r"(odoo\(.*?\))$", val)
                        if len(result):
                            new_val = eval(result[0][5: -1])
                            if isinstance(new_val, float):
                                new_val = str(new_val).replace('.', ',')
                            elif isinstance(new_val, str):
                                new_val = str(new_val)
                            else:
                                # insert image to cell
                                try:
                                    imgdata = base64.b64decode(new_val)
                                    myio = io.BytesIO(imgdata)
                                    img = openpyxl.drawing.image.Image(myio)
                                    cell = ws1.cell(row=row+1, column=column+1)
                                    img.anchor = cell.coordinate
                                    ws1.add_image(img)
                                    new_val = ''
                                except Exception as error:
                                    _logger.error('Error when trying insert image %s' % error)
                                    new_val = ''
                            ws1.cell(
                                row=row+1, column=column+1).value = re.sub(r"(odoo\(.*?\))$", new_val, val)

        # WRITE DATA
        myio = io.BytesIO()
        wb1.save(myio)
        myio.getvalue()
        path_source = self.excel_path_source

        if self.excel_out_report_type != 'excel':
            if not os.path.isdir(self.excel_path_convert_folder):
                os.makedirs(self.excel_path_convert_folder)

            # WRITE DOCX SOURCE
            with open(path_source, 'wb') as f:
                f.write(myio.getbuffer())

            # CONVERT DOCX TO out_report_type
            def convert_to(folder, source, timeout=None):
                args = [self.excel_path_libreoffice, '--headless',
                        '--convert-to', self.excel_out_report_type, '--outdir', folder, source]
                process = subprocess.run(
                    args, stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=timeout)

            convert_to(self.excel_path_convert_folder, path_source)

            # READ out_report_type file FROM OS
            myio = io.BytesIO()
            with open(path_source.replace('xlsx', self.excel_out_report_type), 'rb') as fin:
                myio = io.BytesIO(fin.read())

            try:
                os.unlink(path_source)
                os.unlink(path_source.replace(
                    'xlsx', self.excel_out_report_type))
            except (OSError, IOError):
                _logger.error(
                    'Error when trying to remove file %s' % path_source)

            return myio.getvalue(), self.excel_out_report_type
        return myio.getvalue(), 'excel'
